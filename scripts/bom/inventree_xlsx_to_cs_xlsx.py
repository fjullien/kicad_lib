import openpyxl
import csv
from openpyxl.styles import PatternFill, Alignment, Border
from openpyxl.styles.colors import Color
import openpyxl_dictreader

bom_lines = []
nb_of_manuf = 0

colors = ["FF8030", "008000", "00FF60", "0020A3"]

HEADER_LINE = 11

def get_number_of_manufacturer():
	nb_of_manuf = 0
	for field in reader.fieldnames:
		if field.startswith("Fabricant_"):
			nb_of_manuf += 1
	return nb_of_manuf

def get_number_of_supplier(manuf_index):
	nb_of_supplier = 0
	for field in reader.fieldnames:
		if field.startswith(f"SKU_{manuf_index}_"):
			nb_of_supplier += 1
	return nb_of_supplier

def get_next_row(row):
	i = ord(row[0])
	i += 1
	return chr(i)

def get_previous_row(row):
	i = ord(row[0])
	i -= 1
	return chr(i)

wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("/tmp/bom.xlsx")
sheet = wb.active

# with open("/tmp/CSBRD22005.csv", newline="") as f:
# 	reader = csv.DictReader(f)

reader = openpyxl_dictreader.DictReader("/tmp/in.xlsx")

for n, line in enumerate(reader):
	current_line = n + HEADER_LINE + 1
	compo_list = []

	# Title
	sheet["D2"] = line["parent_part_ipn"] + "\n" + line["parent_part_name"]

	# Gather Manufacturer and associated suppliers for each line
	for m in range(get_number_of_manufacturer()):
		list_manuf_supplier = []
		if line[f"Fabricant_{m}"] != '':
			component = (line[f"Fabricant_{m}"], line[f"MPN_{m}"])
			list_manuf_supplier.append(component)
			for s in range(get_number_of_supplier(m)):
				supplier = (line[f"Fournisseur_{m}_{s}"], line[f"SKU_{m}_{s}"])
				list_manuf_supplier.append(supplier)
			compo_list.append(list_manuf_supplier)

	# Put LCSC first
	for n, c in enumerate(compo_list):
		# Get every suppliers dans associated SKU
		lcsc_pos = 0
		for s in range(len(c) - 1):
			supplier, ref = c[s + 1]
			if supplier == "LCSC" and (s != 0):
				lcsc_pos = s + 1
		if lcsc_pos != 0:
			c[1], c[lcsc_pos] = c[lcsc_pos], c[1]

	# Write new BOM file
	current_row = "J"
	for m in range(get_number_of_manufacturer()):
		sheet[f"{current_row}{HEADER_LINE}"] = "Manufacturer"
		sheet[f"{current_row}{HEADER_LINE}"].fill = PatternFill(start_color=colors[m], fill_type = "solid")
		sheet[f"{current_row}{HEADER_LINE}"].alignment = Alignment(horizontal = 'center')
		current_row = get_next_row(current_row)
		sheet[f"{current_row}{HEADER_LINE}"] = "MPN"
		sheet[f"{current_row}{HEADER_LINE}"].fill = PatternFill(start_color=colors[m], fill_type = "solid")
		sheet[f"{current_row}{HEADER_LINE}"].alignment = Alignment(horizontal = 'center')
		current_row = get_next_row(current_row)
		if line[f"Fabricant_{m}"] != '':
			for s in range(get_number_of_supplier(m)):
				sheet[f"{current_row}{HEADER_LINE}"] = "Supplier"
				sheet[f"{current_row}{HEADER_LINE}"].fill = PatternFill(start_color=colors[m], fill_type = "solid")
				sheet[f"{current_row}{HEADER_LINE}"].alignment = Alignment(horizontal = 'center')
				current_row = get_next_row(current_row)
				sheet[f"{current_row}{HEADER_LINE}"] = "Supplier Part"
				sheet[f"{current_row}{HEADER_LINE}"].fill = PatternFill(start_color=colors[m], fill_type = "solid")
				sheet[f"{current_row}{HEADER_LINE}"].alignment = Alignment(horizontal = 'center')
				current_row = get_next_row(current_row)

	sheet[f"A{current_line}"] = n + 1
	sheet[f"B{current_line}"] = line["quantity"]
	sheet[f"C{current_line}"] = line["reference"]
	sheet[f"D{current_line}"] = line["part_ipn"]
	sheet[f"E{current_line}"] = line["part_name"]
	sheet[f"F{current_line}"] = "X" if line["note"] == "DNP" else line["note"]
	sheet[f"G{current_line}"] = line["Package"]
	if line["Maximum Temperature"] is None:
		temp = "-"
	else:
		temp = line["Maximum Temperature"] + "°C"
		if temp[0] != "+":
			temp = "+" + temp
	sheet[f"H{current_line}"] = temp
	if line["Minimum Temperature"] is None:
		temp = "-"
	else:
		temp = line["Minimum Temperature"] + "°C"
		if (temp[0] != "-") and (temp[0] != "0"):
			temp = "-" + temp
	sheet[f"I{current_line}"] = temp

	current_row = "J"
	for n, c in enumerate(compo_list):
		first_row = current_row
		# Get manufacurer and reference
		manuf, ref = c[0]

		sheet[f"{current_row}{current_line}"] = "-" if manuf is None else manuf
		sheet[f"{current_row}{current_line}"].alignment = Alignment(horizontal = 'center')
		current_row = get_next_row(current_row)

		sheet[f"{current_row}{current_line}"] = "-" if ref is None else ref
		sheet[f"{current_row}{current_line}"].alignment = Alignment(horizontal = 'center')
		current_row = get_next_row(current_row)

		# Get every suppliers dans associated SKU
		for s in range(len(c) - 1):
			supplier, sku = c[s + 1]
			sheet[f"{current_row}{current_line}"] = "-" if supplier is None else supplier
			sheet[f"{current_row}{current_line}"].alignment = Alignment(horizontal = 'center')
			current_row = get_next_row(current_row)
			sheet[f"{current_row}{current_line}"] = "-" if sku is None else sku
			sheet[f"{current_row}{current_line}"].alignment = Alignment(horizontal = 'center')
			current_row = get_next_row(current_row)

		# Put header on top of each component source
		borderStyle = openpyxl.styles.Side(style = 'thin', color = '000000')
		sheet.merge_cells(f"{first_row}{HEADER_LINE - 1}:{get_previous_row(current_row)}{HEADER_LINE - 1}")
		sheet[f"{first_row}{HEADER_LINE - 1}"].fill = PatternFill(start_color=colors[n], fill_type = "solid")
		sheet[f"{first_row}{HEADER_LINE - 1}"].border = Border(left = borderStyle, right = borderStyle, top = borderStyle, bottom = borderStyle)
		sheet[f"{first_row}{HEADER_LINE - 1}"] = f"Source {n + 1}"
		sheet[f"{first_row}{HEADER_LINE - 1}"].alignment = Alignment(horizontal = 'center')

wb.save("/tmp/out.xlsx")